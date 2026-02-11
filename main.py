import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import os
from openpyxl import load_workbook, Workbook
from openpyxl.chart import PieChart, Reference
from datetime import datetime

# Caminho correto da planilha
base_dir = os.path.dirname(os.path.abspath(__file__))
arquivo = os.path.join(base_dir, "Base Atletas.xlsx")
print("Caminho completo:", arquivo)
print("Existe arquivo?", os.path.exists(arquivo))

wb = load_workbook(arquivo)

# Garantir que existe a aba de pagamentos
if "Pagamentos" not in wb.sheetnames:
    ws_pagamentos = wb.create_sheet("Pagamentos")
    ws_pagamentos.append(["Nome", "Valor", "Data", "Mês/Ano", "Forma de Pagamento", "Observações"])
    # Formatar cabeçalho em negrito
    for cell in ws_pagamentos[1]:
        cell.font = cell.font.copy(bold=True)
else:
    ws_pagamentos = wb["Pagamentos"]

# Ajustar largura das colunas
ws_pagamentos.column_dimensions['A'].width = 25  # Nome
ws_pagamentos.column_dimensions['B'].width = 12  # Valor
ws_pagamentos.column_dimensions['C'].width = 15  # Data
ws_pagamentos.column_dimensions['D'].width = 10  # Mês/Ano
ws_pagamentos.column_dimensions['E'].width = 20  # Forma de Pagamento
ws_pagamentos.column_dimensions['F'].width = 30  # Observações

# Aplicar formatação de número na coluna B (Valor) para todas as células
for row in ws_pagamentos.iter_rows(min_row=2, max_row=ws_pagamentos.max_row, min_col=2, max_col=2):
    for cell in row:
        cell.number_format = '#,##0.00'

# Salvar as formatações
wb.save(arquivo)

# Pegar lista de atletas da aba principal
ws_base = wb["Base Atletas"]
atletas = []
for row in ws_base.iter_rows(min_row=2, values_only=True):
    if row[1]:  # Coluna B (índice 1) tem os nomes
        nome_formatado = str(row[1]).strip().title()  # Padroniza: Primeira maiúscula, resto minúsculo
        atletas.append(nome_formatado)

# Função para registrar pagamento
def registrar_pagamento(nome, valor, data, forma, obs):
    # Padronizar o nome antes de salvar
    nome = nome.strip().title()
    
    try:
        data_obj = datetime.strptime(data, "%d/%m/%Y")
        mes_ano = data_obj.strftime("%m/%Y")
    except:
        mes_ano = ""
    
    # Adicionar linha com os dados
    ws_pagamentos.append([nome, float(valor), data, mes_ano, forma, obs])
    
    # Formatar a célula do valor (coluna B) da última linha adicionada
    ultima_linha = ws_pagamentos.max_row
    celula_valor = ws_pagamentos[f'B{ultima_linha}']
    celula_valor.number_format = '#,##0.00'  # Formato de número com 2 casas decimais
    
    wb.save(arquivo)

# Função para gerar relatório com gráfico
def gerar_relatorio_grafico():
    hoje = datetime.today()
    mes_atual = hoje.month
    ano_atual = hoje.year
    
    pagos = set()
    total = 0
    formas = {"Pix": 0, "Dinheiro": 0, "Cartão": 0, "Outro": 0}
    
    for row in ws_pagamentos.iter_rows(min_row=2, values_only=True):
        nome, valor, data, mes_ano, forma, obs = row
        try:
            # Padronizar nome ao verificar pagamento
            nome_padronizado = str(nome).strip().title() if nome else ""
            
            data_pag = datetime.strptime(data, "%d/%m/%Y")
            if data_pag.month == mes_atual and data_pag.year == ano_atual:
                pagos.add(nome_padronizado)
                total += valor
                if forma in formas:
                    formas[forma] += valor
                else:
                    formas["Outro"] += valor
        except:
            continue
    
    devedores = [a for a in atletas if a not in pagos]
    
    # Criar ou limpar aba Resumo
    if "Resumo" in wb.sheetnames:
        del wb["Resumo"]
    
    ws_resumo = wb.create_sheet("Resumo")
    
    # Título e data
    ws_resumo['A1'] = f"RESUMO DE PAGAMENTOS - {hoje.strftime('%m/%Y')}"
    ws_resumo['A1'].font = ws_resumo['A1'].font.copy(bold=True, size=14)
    
    # Dados gerais
    ws_resumo['A3'] = "Total de Atletas:"
    ws_resumo['B3'] = len(atletas)
    ws_resumo['A4'] = "Pagaram:"
    ws_resumo['B4'] = len(pagos)
    ws_resumo['A5'] = "Devem:"
    ws_resumo['B5'] = len(devedores)
    ws_resumo['A6'] = "Total Arrecadado:"
    ws_resumo['B6'] = total
    ws_resumo['B6'].number_format = 'R$ #,##0.00'  # Formato moeda brasileira
    
    # Dados para o gráfico de pizza (Pagantes vs Devedores)
    ws_resumo['D2'] = "Status"
    ws_resumo['E2'] = "Quantidade"
    ws_resumo['F2'] = "Percentual"
    
    ws_resumo['D3'] = "Pagaram"
    ws_resumo['E3'] = len(pagos)
    ws_resumo['F3'] = f"{(len(pagos)/len(atletas)*100):.1f}%" if len(atletas) > 0 else "0%"
    
    ws_resumo['D4'] = "Devem"
    ws_resumo['E4'] = len(devedores)
    ws_resumo['F4'] = f"{(len(devedores)/len(atletas)*100):.1f}%" if len(atletas) > 0 else "0%"
    
    # Criar gráfico de pizza
    pie = PieChart()
    labels = Reference(ws_resumo, min_col=4, min_row=3, max_row=4)
    data = Reference(ws_resumo, min_col=5, min_row=2, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Percentual de Pagamentos"
    
    # Garantir que a legenda fique à direita
    from openpyxl.chart.legend import Legend
    pie.legend = Legend()
    pie.legend.position = 'r'  # 'r' = right (direita)
    
    # Adicionar porcentagens no gráfico
    from openpyxl.chart.label import DataLabelList
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    
    # Posicionar o gráfico
    ws_resumo.add_chart(pie, "D6")
    
    # Remover linhas de grade da aba Resumo
    ws_resumo.sheet_view.showGridLines = False
    
    # Lista de quem pagou
    ws_resumo['A9'] = "PAGARAM:"
    ws_resumo['A9'].font = ws_resumo['A9'].font.copy(bold=True)
    linha = 10
    for nome in sorted(pagos):
        ws_resumo[f'A{linha}'] = nome
        linha += 1
    
    # Lista de quem deve
    ws_resumo[f'C9'] = "DEVEM:"
    ws_resumo['C9'].font = ws_resumo['C9'].font.copy(bold=True)
    linha = 10
    for nome in sorted(devedores):
        ws_resumo[f'C{linha}'] = nome
        linha += 1
    
    # Detalhamento por forma de pagamento
    col_forma = max(15, linha + 2)
    ws_resumo[f'A{col_forma}'] = "FORMAS DE PAGAMENTO:"
    ws_resumo[f'A{col_forma}'].font = ws_resumo[f'A{col_forma}'].font.copy(bold=True)
    
    col_forma += 1
    ws_resumo[f'A{col_forma}'] = "Forma"
    ws_resumo[f'B{col_forma}'] = "Valor"
    
    for forma, valor in formas.items():
        col_forma += 1
        ws_resumo[f'A{col_forma}'] = forma
        ws_resumo[f'B{col_forma}'] = valor
        ws_resumo[f'B{col_forma}'].number_format = 'R$ #,##0.00'  # Formato moeda
    
    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 25
    ws_resumo.column_dimensions['B'].width = 15
    ws_resumo.column_dimensions['C'].width = 25
    ws_resumo.column_dimensions['D'].width = 15
    ws_resumo.column_dimensions['E'].width = 15
    ws_resumo.column_dimensions['F'].width = 15
    
    # Salvar planilha
    wb.save(arquivo)
    
    return len(pagos), len(devedores), total

# Funções da interface
def btn_registrar_pagamento():
    nome = combo_nome.get()
    valor = entry_valor.get()
    data = entry_data.get()
    forma = combo_forma.get()
    obs = entry_obs.get()
    
    if not nome or not valor or not data or not forma:
        text_status.delete(1.0, tk.END)
        text_status.insert(tk.END, "⚠️ Preencha todos os campos obrigatórios!")
    else:
        try:
            registrar_pagamento(nome, valor, data, forma, obs)
            text_status.delete(1.0, tk.END)
            text_status.insert(tk.END, f"✅ Pagamento registrado para {nome.title()}!\n")
            text_status.insert(tk.END, f"💰 Valor: R$ {float(valor):.2f}\n")
            text_status.insert(tk.END, f"📅 Data: {data}\n")
            text_status.insert(tk.END, f"💳 Forma: {forma}\n")
            text_status.insert(tk.END, f"\n✔️ Salvo na aba 'Pagamentos' do Excel!")
            
            # Limpar campos após registrar
            entry_valor.delete(0, tk.END)
            entry_obs.delete(0, tk.END)
        except Exception as e:
            text_status.delete(1.0, tk.END)
            text_status.insert(tk.END, f"❌ Erro: {e}")

def btn_gerar_relatorio():
    try:
        pagos, devedores, total = gerar_relatorio_grafico()
        
        text_status.delete(1.0, tk.END)
        text_status.insert(tk.END, f"✅ RELATÓRIO GERADO COM SUCESSO!\n\n")
        text_status.insert(tk.END, f"📊 Uma nova aba 'Resumo' foi criada na planilha Excel\n\n")
        text_status.insert(tk.END, f"📈 RESUMO:\n")
        text_status.insert(tk.END, f"   • Total de atletas: {len(atletas)}\n")
        text_status.insert(tk.END, f"   • Pagaram: {pagos} ({(pagos/len(atletas)*100):.1f}%)\n")
        text_status.insert(tk.END, f"   • Devem: {devedores} ({(devedores/len(atletas)*100):.1f}%)\n")
        text_status.insert(tk.END, f"   • Total arrecadado: R$ {total:.2f}\n\n")
        text_status.insert(tk.END, f"📂 Abra o arquivo Excel para ver o gráfico completo!")
        
    except Exception as e:
        text_status.delete(1.0, tk.END)
        text_status.insert(tk.END, f"❌ Erro ao gerar relatório: {e}")
        messagebox.showerror("Erro", f"Não foi possível gerar o relatório:\n{e}")

# Criar janela principal
root = tk.Tk()
root.title("Mensalidade Futebol - Veteranos Vila Avaí")
root.geometry("495x520")
root.resizable(False, False)

# Configurar estilo
style = ttk.Style()
style.theme_use('clam')

# Criar frame principal
frame_principal = ttk.Frame(root, padding="15")
frame_principal.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Título
titulo = ttk.Label(frame_principal, text="⚽ CONTROLE DE PAGAMENTOS", font=('Arial', 14, 'bold'))
titulo.grid(row=0, column=0, columnspan=2, pady=10)

# Campos de entrada
ttk.Label(frame_principal, text="Selecione o atleta:", font=('Arial', 10)).grid(row=1, column=0, sticky=tk.W, pady=5)
combo_nome = ttk.Combobox(frame_principal, values=atletas, width=37, font=('Arial', 10))
combo_nome.grid(row=1, column=1, sticky=tk.W, pady=5)

ttk.Label(frame_principal, text="Valor pago:", font=('Arial', 10)).grid(row=2, column=0, sticky=tk.W, pady=5)
entry_valor = ttk.Entry(frame_principal, width=17, font=('Arial', 10))
entry_valor.grid(row=2, column=1, sticky=tk.W, pady=5)

ttk.Label(frame_principal, text="Data do pagamento:", font=('Arial', 10)).grid(row=3, column=0, sticky=tk.W, pady=5)
entry_data = ttk.Entry(frame_principal, width=17, font=('Arial', 10))
entry_data.insert(0, datetime.today().strftime("%d/%m/%Y"))
entry_data.grid(row=3, column=1, sticky=tk.W, pady=5)

ttk.Label(frame_principal, text="Forma de pagamento:", font=('Arial', 10)).grid(row=4, column=0, sticky=tk.W, pady=5)
combo_forma = ttk.Combobox(frame_principal, values=["Pix", "Dinheiro", "Cartão", "Outro"], width=17, font=('Arial', 10))
combo_forma.grid(row=4, column=1, sticky=tk.W, pady=5)

ttk.Label(frame_principal, text="Observações:", font=('Arial', 10)).grid(row=5, column=0, sticky=tk.W, pady=5)
entry_obs = ttk.Entry(frame_principal, width=37, font=('Arial', 10))
entry_obs.grid(row=5, column=1, sticky=tk.W, pady=5)

# Frame de botões - APENAS 2 BOTÕES
frame_botoes = ttk.Frame(frame_principal)
frame_botoes.grid(row=6, column=0, columnspan=2, pady=15)

btn_registrar = ttk.Button(frame_botoes, text="💾 Registrar Pagamento", command=btn_registrar_pagamento, width=25)
btn_registrar.pack(side=tk.LEFT, padx=10)

btn_relatorio = ttk.Button(frame_botoes, text="📊 Gerar Relatório", command=btn_gerar_relatorio, width=25)
btn_relatorio.pack(side=tk.LEFT, padx=10)

# Área de status/relatório
ttk.Label(frame_principal, text="Status:", font=('Arial', 10, 'bold')).grid(row=7, column=0, columnspan=2, sticky=tk.W, pady=5)
text_status = scrolledtext.ScrolledText(frame_principal, width=63, height=10, wrap=tk.WORD, font=('Consolas', 9))
text_status.grid(row=8, column=0, columnspan=2, pady=5)

# Mensagem inicial
text_status.insert(tk.END, "✅ Sistema iniciado com TKINTER (100% gratuito!)\n")
text_status.insert(tk.END, f"📁 Planilha: Base Atletas.xlsx\n")
text_status.insert(tk.END, f"👥 Total de atletas: {len(atletas)}\n\n")
text_status.insert(tk.END, "Aguardando operação...")

# Botão Sair no rodapé
btn_sair = ttk.Button(frame_principal, text="❌ Sair", command=root.quit, width=15)
btn_sair.grid(row=9, column=0, columnspan=2, pady=10)

# Iniciar aplicação
root.mainloop()
